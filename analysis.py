from aip import AipNlp
from openpyxl import load_workbook
import time
from openpyxl import Workbook

# 创建你的 AppID
APP_ID = '41857222'
API_KEY = 'Y7Ry5AyVSSU6LZVIyz8Z1Par'
SECRET_KEY = 'nptEz49nFB403S5bEsxqG9eYzl0OwkGs'
client = AipNlp(APP_ID, API_KEY, SECRET_KEY)

# 加载 Excel 文件
wb = load_workbook("lhb.xlsx")
ws = wb.active

# 获取所有评论
comments_column = None
for cell in ws[1]:
    if cell.value == "评论内容":
        comments_column = cell.column
        break

if comments_column is None:
    print("找不到评论内容列")
else:
    result_data = [["评论内容", "情感积极", "情感消极", "置信度"]]

    for row in range(2, ws.max_row + 1):
        comment = ws.cell(column=comments_column, row=row).value

        # 使用百度API进行情感分析
        result = client.sentimentClassify(comment)

        if 'items' in result:
            emotion = result['items'][0]
            positive_prob = round(emotion['positive_prob'] * 100, 2)
            negative_prob = round(emotion['negative_prob'] * 100, 2)
            confidence = round(emotion['confidence'] * 100, 2)
        else:
            positive_prob = 0
            negative_prob = 0
            confidence = 0

        result_data.append([comment, positive_prob, negative_prob, confidence])

        # 在每次循环后等待一段时间
        time.sleep(1)  # 这里等待 1 秒

    # 创建一个新的 Excel 文件来保存结果
    result_wb = Workbook()
    result_ws = result_wb.active

    for row in result_data:
        result_ws.append(row)

    result_wb.save("emotionlhb.xlsx")
    print("情感分析结果已保存到 emotion.xlsx")
